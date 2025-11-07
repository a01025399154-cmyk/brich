"""
엑셀 파일 생성 모듈
시작일-종료일_채널별로 분리된 파일 생성
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, numbers
from typing import List


def generate_upload_files(df: pd.DataFrame, output_dir: str, file_prefix: str = "상품") -> List[str]:
    """
    채널별로 분리된 업로드 파일들 생성
    
    Args:
        df: 출력 데이터 DataFrame (시작일, 종료일, 채널명 포함)
        output_dir: 출력 디렉토리
        file_prefix: 파일명 prefix ("상품" 또는 "브랜드")
    
    Returns:
        생성된 파일 경로 리스트
    """
    # 출력 디렉토리 생성
    os.makedirs(output_dir, exist_ok=True)
    
    # (시작일, 종료일, 채널명)으로 그룹화
    grouped = df.groupby(['시작일', '종료일', '채널명'])
    
    generated_files = []
    
    for (start_date, end_date, channel_name), group_df in grouped:
        # 파일명 생성
        start_str = start_date.strftime('%y%m%d') if pd.notna(start_date) else '000000'
        end_str = end_date.strftime('%y%m%d') if pd.notna(end_date) else '000000'
        item_count = len(group_df)
        
        filename = f"{start_str}-{end_str}_{file_prefix}_{channel_name}_{item_count}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # 엑셀에 저장할 컬럼 선택 (시작일, 종료일, 채널명 제외)
        columns_to_save = [col for col in group_df.columns if col not in ['시작일', '종료일', '채널명']]
        df_to_save = group_df[columns_to_save].copy()
        
        # 상품번호/브랜드번호를 문자열로 변환
        number_col = '상품번호' if '상품번호' in df_to_save.columns else '브랜드번호'
        if number_col in df_to_save.columns:
            # NaN이 아닌 값만 정수로 변환 후 문자열로
            df_to_save[number_col] = df_to_save[number_col].apply(
                lambda x: str(int(float(x))) if pd.notna(x) else ''
            )
        
        # 할인 컬럼만 정수로 변환 (내부할인, 연동할인, 외부할인가, 할인, 할인2)
        discount_value_cols = ['내부할인', '연동할인', '외부할인가', '할인', '할인2']
        for col in discount_value_cols:
            if col in df_to_save.columns:
                df_to_save[col] = df_to_save[col].apply(
                    lambda x: int(float(x)) if pd.notna(x) and str(x) != '' else 0
                )
        
        # 분담율은 문자열로 변환
        rate_cols = ['채널분담율', '브리치분담율', '입점사분담율']
        for col in rate_cols:
            if col in df_to_save.columns:
                df_to_save[col] = df_to_save[col].apply(
                    lambda x: str(int(float(x))) if pd.notna(x) and str(x) != '' else '0'
                )
        
        # 엑셀 파일 생성
        df_to_save.to_excel(filepath, index=False, engine='openpyxl')
        
        # ✅ FIX: 포맷팅 - 브랜드 프로모션 호환을 위해 모든 셀을 텍스트로
        wb = load_workbook(filepath)
        ws = wb.active
        
        # 헤더 스타일 - 모든 헤더를 텍스트 형식으로
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = '@'
        
        # 모든 데이터 셀을 텍스트 형식으로 설정
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row, col)
                if cell.value is not None:
                    # 브랜드번호/상품번호는 문자열로 확실히 변환
                    col_name = ws.cell(1, col).value
                    if col_name in ['브랜드번호', '상품번호']:
                        try:
                            cell.value = str(int(float(cell.value)))
                        except:
                            cell.value = str(cell.value)
                
                # ✅ FIX: 할인 값 컬럼은 숫자 포맷(General)으로 유지
                col_name = ws.cell(1, col).value
                if col_name in ['할인', '내부할인', '연동할인', '외부할인가']:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = 'General'
                    else:
                        cell.number_format = '@'
                else:
                    cell.number_format = '@'
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 저장
        wb.save(filepath)
        
        # 비플로우 양식 맞추기: 빈 행/열 추가 및 포맷 설정
        wb = load_workbook(filepath)
        ws = wb.active
        
        # ✅ FIX: 시트명을 '시트1'로 변경 (비플로우 호환)
        ws.title = '시트1'
        
        # ✅ FIX: 최소 999행 x 40열로 확장 (수동 업로드 파일과 동일)
        min_rows = 999
        min_cols = 40
        
        # 할인 값 컬럼 인덱스 찾기
        discount_cols = []
        for col_idx in range(1, ws.max_column + 1):
            col_name = ws.cell(1, col_idx).value
            if col_name in ['할인', '내부할인', '연동할인', '외부할인가']:
                discount_cols.append(col_idx)
        
        # 모든 셀에 텍스트 포맷 적용 (빈 셀 포함)
        for row_idx in range(1, min_rows + 1):
            for col_idx in range(1, min_cols + 1):
                cell = ws.cell(row_idx, col_idx)
                # ✅ FIX: 할인 값 컬럼은 General 포맷 유지 (빈 셀 포함)
                if col_idx in discount_cols and row_idx > 1:
                    cell.number_format = 'General'
                else:
                    cell.number_format = '@'
        
        wb.save(filepath)
        
        generated_files.append(filepath)
        print(f"  ✓ 생성: {filename}")
    
    return generated_files


if __name__ == "__main__":
    # 테스트 - 상품 프로모션
    df_test_product = pd.DataFrame({
        '시작일': [pd.Timestamp('2025-11-01')] * 4,
        '종료일': [pd.Timestamp('2025-12-05')] * 4,
        '채널명': ['SSG', 'SSG', '쿠팡', '쿠팡'],
        '상품번호': [1000614610607, 1000614610608, 200012345678, 200012345679],
        '내부할인타입': ['P', 'P', 'P', 'P'],
        '내부할인': [17, 17, 17, 17],
        '연동할인타입': ['', '', '', ''],
        '연동할인': ['', '', '', ''],
        '외부할인타입': ['', '', '', ''],
        '외부할인가': ['', '', '', ''],
        '채널분담율': ['', '', '', ''],
        '브리치분담율': ['', '', '', ''],
        '입점사분담율': ['', '', '', '']
    })
    
    output_files = generate_upload_files(df_test_product, '/home/claude/test_output', '상품')
    print(f"\n상품 테스트 완료: {len(output_files)}개 파일 생성")
    
    # 테스트 - 브랜드 프로모션
    df_test_brand = pd.DataFrame({
        '시작일': [pd.Timestamp('2025-10-01')] * 3,
        '종료일': [pd.Timestamp('2025-10-31')] * 3,
        '채널명': ['SSG', 'SSG', '쿠팡'],
        '브랜드번호': [1838, 3197, 1838],
        '할인타입': ['P', 'P', 'P'],
        '할인': [15, 15, 15],
        '채널분담율': [0, 0, 0],
        '브리치분담율': [0, 0, 0],
        '입점사분담율': [100, 100, 100]
    })
    
    output_files = generate_upload_files(df_test_brand, '/home/claude/test_output', '브랜드')
    print(f"\n브랜드 테스트 완료: {len(output_files)}개 파일 생성")