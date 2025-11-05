"""
엑셀 파일 생성 모듈
시작일-종료일_채널별로 분리된 파일 생성
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from collections import defaultdict

def generate_upload_files(df: pd.DataFrame, output_dir: str) -> list:
    """
    채널별로 분리된 product_promotion_upload 파일들 생성
    
    Args:
        df: 출력 데이터 DataFrame (시작일, 종료일, 채널명 포함)
        output_dir: 출력 디렉토리
    
    Returns:
        생성된 파일 경로 리스트
    """
    # 출력 디렉토리 생성
    os.makedirs(output_dir, exist_ok=True)
    
    # (시작일, 종료일, 채널명)으로 그룹화
    grouped = df.groupby(['시작일', '종료일', '채널명'])
    
    generated_files = []
    
    for (start_date, end_date, channel_name), group_df in grouped:
        # 파일명 생성
        start_str = start_date.strftime('%y%m%d') if pd.notna(start_date) else '000000'
        end_str = end_date.strftime('%y%m%d') if pd.notna(end_date) else '000000'
        product_count = len(group_df)
        
        filename = f"{start_str}-{end_str}_상품_{channel_name}_{product_count}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # 엑셀에 저장할 컬럼만 선택
        output_columns = [
            '상품번호',
            '내부할인타입',
            '내부할인',
            '연동할인타입',
            '연동할인',
            '외부할인타입',
            '외부할인가',
            '채널분담율',
            '브리치분담율',
            '입점사분담율'
        ]
        
        df_to_save = group_df[output_columns].copy()
        
        # 엑셀 파일 생성
        df_to_save.to_excel(filepath, index=False, engine='openpyxl')
        
        # 포맷팅
        wb = load_workbook(filepath)
        ws = wb.active
        
        # 헤더 스타일
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 저장
        wb.save(filepath)
        
        generated_files.append(filepath)
        print(f"  ✓ 생성: {filename}")
    
    return generated_files


if __name__ == "__main__":
    # 테스트
    df_test = pd.DataFrame({
        '시작일': [pd.Timestamp('2025-11-01')] * 6,
        '종료일': [pd.Timestamp('2025-12-05')] * 6,
        '채널명': ['SSG', 'SSG', '쿠팡', '쿠팡', '지마켓', '지마켓'],
        '상품번호': ['1000614610607', '1000614610608', '200012345678', '200012345679', '123456789', '123456790'],
        '내부할인타입': ['P', 'P', 'P', 'P', 'P', 'P'],
        '내부할인': [17, 17, 17, 17, 17, 17],
        '연동할인타입': ['', '', '', '', '', ''],
        '연동할인': ['', '', '', '', '', ''],
        '외부할인타입': ['', '', '', '', '', ''],
        '외부할인가': ['', '', '', '', '', ''],
        '채널분담율': ['', '', '', '', '', ''],
        '브리치분담율': ['', '', '', '', '', ''],
        '입점사분담율': ['', '', '', '', '', '']
    })
    
    output_files = generate_upload_files(df_test, '/home/claude/test_output')
    print(f"\n테스트 완료: {len(output_files)}개 파일 생성")